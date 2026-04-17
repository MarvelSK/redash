from __future__ import annotations

import argparse
import logging
import os
import re
import shutil
import sys
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import mysql.connector
from mysql.connector import MySQLConnection
from openpyxl import load_workbook
import xlrd


LOGGER = logging.getLogger("people_counter_etl")

DEFAULT_MIN_ZIP_SIZE_BYTES = 5 * 1024
DEFAULT_ROOT_FOLDER = Path(r"C:\Users\ch800\Documents\KPI_COCKPIT")
DEFAULT_SOURCE_FOLDER = Path(r"\\chwinapp04\Transfer_People_Counter")


@dataclass(frozen=True)
class AppConfig:
    source_folder: Path
    root_folder: Path
    incoming_folder: Path
    work_folder: Path
    archive_folder: Path
    min_zip_size_bytes: int
    mysql_host: str
    mysql_port: int
    mysql_user: str
    mysql_password: str
    mysql_database: str


def env_str(name: str, default: str) -> str:
    value = os.getenv(name)
    return value if value not in (None, "") else default


def env_int(name: str, default: int) -> int:
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return int(value)


def load_config() -> AppConfig:
    root_folder = Path(env_str("PEOPLE_COUNTER_ROOT_FOLDER", str(DEFAULT_ROOT_FOLDER)))
    source_folder = Path(env_str("PEOPLE_COUNTER_SOURCE_FOLDER", str(DEFAULT_SOURCE_FOLDER)))
    incoming_folder = root_folder / "incoming"
    work_folder = root_folder / "work"
    archive_folder = root_folder / "archive"

    return AppConfig(
        source_folder=source_folder,
        root_folder=root_folder,
        incoming_folder=incoming_folder,
        work_folder=work_folder,
        archive_folder=archive_folder,
        min_zip_size_bytes=env_int("PEOPLE_COUNTER_MIN_ZIP_SIZE_BYTES", DEFAULT_MIN_ZIP_SIZE_BYTES),
        mysql_host=env_str("MYSQL_HOST", "127.0.0.1"),
        mysql_port=env_int("MYSQL_PORT", 3306),
        mysql_user=env_str("MYSQL_USER", "people_counter"),
        mysql_password=env_str("MYSQL_PASSWORD", "change_me"),
        mysql_database=env_str("MYSQL_DATABASE", "people_counter"),
    )


def configure_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )


def remove_diacritics(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_text.lower()


def get_excel_cell(sheet: object, row_index: int, column_index: int) -> object:
    if hasattr(sheet, "cell"):
        cell = sheet.cell(row=row_index, column=column_index)
        if hasattr(cell, "value"):
            return cell.value
    return sheet.cell_value(row_index - 1, column_index - 1)


class WorkbookAdapter:
    def __init__(self, excel_path: Path) -> None:
        self._workbook = None
        self._mode = "openpyxl"

        if excel_path.suffix.lower() == ".xls":
            self._workbook = xlrd.open_workbook(excel_path)
            self._mode = "xlrd"
        else:
            self._workbook = load_workbook(excel_path, data_only=True, read_only=True)

    def get_sheet(self, index: int) -> object:
        if self._mode == "xlrd":
            return self._workbook.sheet_by_index(index)
        return self._workbook.worksheets[index]

    def close(self) -> None:
        if self._mode == "openpyxl" and self._workbook is not None:
            self._workbook.close()


def safe_int(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return 0
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return 0


def is_probable_age_bucket(label: str) -> bool:
    normalized = remove_diacritics(label)
    if not normalized:
        return False

    # Never treat time/date labels as age buckets.
    if ":" in normalized:
        return False
    if re.search(r"\b\d{1,2}/\d{1,2}\b", normalized):
        return False

    # Skip obvious non-age rows.
    if normalized in {"mannlich", "weiblich", "nicht erkannt", "gesamt", "summe", "total"}:
        return False

    # Skip time labels like 11:00.
    if re.match(r"^\d{1,2}:\d{2}$", normalized):
        return False

    # Typical age labels in reports are either text-based buckets (Jugendliche...)
    # or numeric ranges (0-9, 70+).
    if any(token in normalized for token in ("jugend", "kinder", "senior", "alter", "teen")):
        return True

    if re.match(r"^\d+\s*-\s*\d+\+?$", normalized) or re.match(r"^\d+\+$", normalized):
        return True

    # Accept labels like "15-17 Jahre" by looking for range-like tokens.
    if re.search(r"\d+\s*-\s*\d+", normalized):
        return True

    return False


def parse_age_items(sheet: object) -> list[tuple[str, int]]:
    # Scan a wider area and detect labels in any column with numeric value nearby.
    found: list[tuple[str, int]] = []
    seen_labels: set[str] = set()

    for row_index in range(1, 121):
        row_values = [get_excel_cell(sheet, row_index, col_index) for col_index in range(1, 9)]

        for col_index, raw_label in enumerate(row_values, start=1):
            label = str(raw_label or "").strip()
            if not label:
                continue
            if not is_probable_age_bucket(label):
                continue

            # Find first numeric value to the right of the label (up to 3 columns).
            count = 0
            for offset in (1, 2, 3):
                next_col = col_index + offset
                if next_col > 8:
                    break
                value = get_excel_cell(sheet, row_index, next_col)
                if isinstance(value, (int, float)):
                    count = int(value)
                    break
                parsed = safe_int(value)
                if parsed != 0:
                    count = parsed
                    break

            if label not in seen_labels:
                found.append((label, count))
                seen_labels.add(label)

    return found


def summarize_rows_for_debug(sheet: object, max_rows: int = 80) -> str:
    samples: list[str] = []
    for row_index in range(1, max_rows + 1):
        values = [get_excel_cell(sheet, row_index, col_index) for col_index in range(1, 8)]
        if any(v not in (None, "") for v in values):
            compact = " | ".join(str(v) if v is not None else "" for v in values)
            samples.append(f"r{row_index}: {compact}")
        if len(samples) >= 12:
            break
    return " ; ".join(samples)


def parse_report_date(raw_value: object) -> date:
    text = str(raw_value or "").strip()
    match = re.search(r"(\d{4}/\d{2}/\d{2})", text)
    if not match:
        return date.today()
    return datetime.strptime(match.group(1), "%Y/%m/%d").date()


def ensure_folders(config: AppConfig) -> None:
    for folder in (config.root_folder, config.incoming_folder, config.work_folder, config.archive_folder):
        folder.mkdir(parents=True, exist_ok=True)


def get_connection(config: AppConfig) -> MySQLConnection:
    return mysql.connector.connect(
        host=config.mysql_host,
        port=config.mysql_port,
        user=config.mysql_user,
        password=config.mysql_password,
        database=config.mysql_database,
        autocommit=False,
    )


def move_zips_from_share(config: AppConfig) -> None:
    if not config.source_folder.exists():
        LOGGER.warning("Source folder does not exist: %s", config.source_folder)
        return

    for path in config.source_folder.iterdir():
        if path.suffix.lower() != ".zip":
            continue
        if path.stat().st_size <= config.min_zip_size_bytes:
            LOGGER.info("Skipping small zip: %s", path.name)
            continue
        target = config.incoming_folder / path.name
        if target.exists():
            target.unlink()
        shutil.move(str(path), str(target))
        LOGGER.info("Moved zip to incoming: %s", target)


def iter_incoming_zips(config: AppConfig) -> Iterable[Path]:
    if not config.incoming_folder.exists():
        return []
    return sorted(path for path in config.incoming_folder.iterdir() if path.suffix.lower() == ".zip")


def unzip_to_folder(zip_path: Path, target_folder: Path) -> bool:
    target_folder.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(zip_path, "r") as archive:
            archive.extractall(target_folder)
        return any(target_folder.iterdir())
    except zipfile.BadZipFile:
        return False


def fetch_active_shops(connection: MySQLConnection) -> list[dict[str, str]]:
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT shop_id, shop_name FROM cfg_shops WHERE active = 1")
    rows = cursor.fetchall()
    cursor.close()
    return rows


def resolve_shop_from_zip(connection: MySQLConnection, zip_name: str) -> str | None:
    zip_norm = remove_diacritics(zip_name)
    for row in fetch_active_shops(connection):
        shop_norm = remove_diacritics(str(row["shop_name"]))
        if shop_norm and shop_norm in zip_norm:
            return str(row["shop_id"])
    return None


def get_shop_name(connection: MySQLConnection, shop_id: str) -> str | None:
    cursor = connection.cursor()
    cursor.execute("SELECT shop_name FROM cfg_shops WHERE shop_id = %s", (shop_id,))
    row = cursor.fetchone()
    cursor.close()
    return str(row[0]) if row else None


def find_excel_by_shop(folder_path: Path, shop_name: str, shop_id: str) -> Path | None:
    excel_files = [path for path in folder_path.iterdir() if path.is_file() and path.suffix.lower() in {".xls", ".xlsx", ".xlsm"}]
    if len(excel_files) == 1:
        return excel_files[0]

    want_name = remove_diacritics(shop_name or "")
    want_id = remove_diacritics(shop_id or "")

    for path in excel_files:
        file_norm = remove_diacritics(path.name)
        if (want_name and want_name in file_norm) or (want_id and want_id in file_norm):
            return path

    for child in folder_path.iterdir():
        if child.is_dir():
            found = find_excel_by_shop(child, shop_name, shop_id)
            if found:
                return found
    return None


def log_import(
    connection: MySQLConnection,
    zip_name: str,
    shop_id: str | None,
    status: str,
    message: str,
    kpi_date: date | None = None,
) -> None:
    cursor = connection.cursor()
    cursor.execute(
        """
        INSERT INTO etl_import_log (run_timestamp, shop_id, kpi_date, status, message)
        VALUES (%s, %s, %s, %s, %s)
        """,
        (
            datetime.now(),
            shop_id,
            kpi_date,
            status,
            f"{message} | ZIP={zip_name}",
        ),
    )
    cursor.close()


def upsert_daily_overview(
    connection: MySQLConnection,
    shop_id: str,
    kpi_date: date,
    visitors_total: int,
    visitors_male: int,
    visitors_female: int,
    visitors_unknown: int,
) -> None:
    cursor = connection.cursor()
    cursor.execute(
        """
        INSERT INTO kpi_daily_overview
            (shop_id, kpi_date, visitors_total, visitors_male, visitors_female, visitors_unknown, sales_net)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            visitors_total = VALUES(visitors_total),
            visitors_male = VALUES(visitors_male),
            visitors_female = VALUES(visitors_female),
            visitors_unknown = VALUES(visitors_unknown),
            sales_net = VALUES(sales_net)
        """,
        (shop_id, kpi_date, visitors_total, visitors_male, visitors_female, visitors_unknown, None),
    )
    cursor.close()


def replace_age_distribution(connection: MySQLConnection, shop_id: str, kpi_date: date, items: list[tuple[str, int]]) -> None:
    cursor = connection.cursor()
    cursor.execute(
        "DELETE FROM kpi_age_distribution WHERE shop_id = %s AND kpi_date = %s",
        (shop_id, kpi_date),
    )
    cursor.executemany(
        """
        INSERT INTO kpi_age_distribution (shop_id, kpi_date, age_bucket, visitors_count)
        VALUES (%s, %s, %s, %s)
        """,
        [(shop_id, kpi_date, age_bucket, visitors_count) for age_bucket, visitors_count in items],
    )
    cursor.close()


def replace_hourly_visitors(
    connection: MySQLConnection,
    shop_id: str,
    kpi_date: date,
    items: list[tuple[str, int, int, int, int]],
) -> None:
    cursor = connection.cursor()
    cursor.execute(
        "DELETE FROM kpi_hourly_visitors WHERE shop_id = %s AND kpi_date = %s",
        (shop_id, kpi_date),
    )
    cursor.executemany(
        """
        INSERT INTO kpi_hourly_visitors
            (shop_id, kpi_date, visit_hour, visitors_male, visitors_female, visitors_unknown, visitors_total)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        [(shop_id, kpi_date, visit_hour, male, female, unknown, total) for visit_hour, male, female, unknown, total in items],
    )
    cursor.close()


def parse_and_write(connection: MySQLConnection, excel_path: Path, shop_id: str) -> date:
    workbook = WorkbookAdapter(excel_path)
    try:
        sheet1 = workbook.get_sheet(0)
        sheet2 = workbook.get_sheet(1)

        kpi_date = parse_report_date(get_excel_cell(sheet1, 6, 1))

        male = safe_int(get_excel_cell(sheet2, 11, 1))
        female = safe_int(get_excel_cell(sheet2, 11, 3))
        unknown = safe_int(get_excel_cell(sheet2, 11, 5))
        total = male + female + unknown

        upsert_daily_overview(connection, shop_id, kpi_date, total, male, female, unknown)

        age_items = parse_age_items(sheet2)

        if age_items:
            replace_age_distribution(connection, shop_id, kpi_date, age_items)
        else:
            LOGGER.warning(
                "No age rows parsed for shop=%s date=%s file=%s; sample=%s; keeping any existing age rows unchanged",
                shop_id,
                kpi_date,
                excel_path,
                summarize_rows_for_debug(sheet2),
            )

        hourly_items: list[tuple[str, int, int, int, int]] = []
        for row_index in range(33, 49):
            hour_label = str(get_excel_cell(sheet2, row_index, 2) or "").strip()
            if not hour_label:
                continue
            male_hour = safe_int(get_excel_cell(sheet2, row_index, 3))
            female_hour = safe_int(get_excel_cell(sheet2, row_index, 5))
            unknown_hour = safe_int(get_excel_cell(sheet2, row_index, 7))
            total_hour = male_hour + female_hour + unknown_hour
            hourly_items.append((hour_label, male_hour, female_hour, unknown_hour, total_hour))

        replace_hourly_visitors(connection, shop_id, kpi_date, hourly_items)
        return kpi_date
    finally:
        workbook.close()


def process_one_zip(connection: MySQLConnection, config: AppConfig, zip_path: Path) -> None:
    zip_name = zip_path.name
    shop_id: str | None = None
    kpi_date: date | None = None

    try:
        if not zip_path.exists():
            log_import(connection, zip_name, shop_id, "ERROR", f"ZIP file not found: {zip_path}")
            connection.commit()
            return

        output_folder = config.work_folder / zip_path.stem
        output_folder.mkdir(parents=True, exist_ok=True)

        if not unzip_to_folder(zip_path, output_folder):
            log_import(connection, zip_name, shop_id, "ERROR", f"UNZIP failed | folder={output_folder}")
            connection.commit()
            return

        if not any(output_folder.iterdir()):
            log_import(connection, zip_name, shop_id, "ERROR", f"WORK folder empty after UNZIP | folder={output_folder}")
            connection.commit()
            return

        shop_id = resolve_shop_from_zip(connection, zip_name)
        if not shop_id:
            log_import(connection, zip_name, shop_id, "ERROR", f"Shop not resolved from cfg_shops.shop_name | ZIP={zip_name}")
            connection.commit()
            return

        shop_name = get_shop_name(connection, shop_id)
        if not shop_name:
            log_import(connection, zip_name, shop_id, "ERROR", f"shop_name not found in cfg_shops | shop_id={shop_id}")
            connection.commit()
            return

        excel_path = find_excel_by_shop(output_folder, shop_name, shop_id)
        if not excel_path:
            log_import(
                connection,
                zip_name,
                shop_id,
                "ERROR",
                f"Excel not found | shop_name={shop_name} | shop_id={shop_id} | folder={output_folder}",
            )
            connection.commit()
            return

        kpi_date = parse_and_write(connection, excel_path, shop_id)

        archive_target = config.archive_folder / zip_name
        if archive_target.exists():
            archive_target.unlink()
        shutil.move(str(zip_path), str(archive_target))

        log_import(connection, zip_name, shop_id, "OK", f"Processed | excel={excel_path}", kpi_date)
        connection.commit()
    except Exception as exc:
        connection.rollback()
        try:
            log_import(connection, zip_name, shop_id, "ERROR", str(exc), kpi_date)
            connection.commit()
        except Exception:
            connection.rollback()
        raise


def process_incoming_zips(connection: MySQLConnection, config: AppConfig, single_zip: bool) -> int:
    processed = 0
    for zip_path in iter_incoming_zips(config):
        process_one_zip(connection, config, zip_path)
        processed += 1
        if single_zip:
            break
    return processed


def run_etl(single_zip: bool, verbose: bool) -> int:
    configure_logging(verbose)
    config = load_config()
    ensure_folders(config)

    move_zips_from_share(config)

    connection = get_connection(config)
    try:
        processed = process_incoming_zips(connection, config, single_zip=single_zip)
        LOGGER.info("ETL finished. Processed ZIP files: %s", processed)
        return 0
    finally:
        connection.close()


def seed_database(config: AppConfig, schema_path: Path, seed_path: Path) -> int:
    connection = mysql.connector.connect(
        host=config.mysql_host,
        port=config.mysql_port,
        user=config.mysql_user,
        password=config.mysql_password,
        autocommit=True,
    )
    try:
        cursor = connection.cursor()
        for script_path in (schema_path, seed_path):
            sql = script_path.read_text(encoding="utf-8")
            for statement in [part.strip() for part in sql.split(";") if part.strip()]:
                cursor.execute(statement)
        cursor.close()
        return 0
    finally:
        connection.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="People Counter ETL to MySQL")
    subparsers = parser.add_subparsers(dest="command")

    run_parser = subparsers.add_parser("run", help="Run ETL")
    run_parser.add_argument("--single-zip", action="store_true", help="Process only one ZIP file")
    run_parser.add_argument("--verbose", action="store_true", help="Enable debug logging")

    seed_parser = subparsers.add_parser("seed", help="Create schema and seed baseline data")
    seed_parser.add_argument("--schema", default="mysql_schema.sql", help="Path to schema SQL file")
    seed_parser.add_argument("--seed", default="mysql_seed.sql", help="Path to seed SQL file")

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.command == "seed":
        config = load_config()
        return seed_database(config, Path(args.schema), Path(args.seed))

    if args.command == "run":
        return run_etl(single_zip=args.single_zip, verbose=args.verbose)

    parser.print_help()
    return 1


if __name__ == "__main__":
    sys.exit(main())